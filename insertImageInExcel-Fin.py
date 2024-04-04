from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

xlfile_name = "currentRes.xlsx"
png_file_name = "currentPareto.png"
png2 = "currentPareto1.png"

# Load the existing workbook
wb = load_workbook(xlfile_name)

# Add a new worksheet to the workbook
new_ws = wb.create_sheet(title="Sheet3")

# Select the active worksheet (Sheet1)
ws = wb.active

# Load the PNG images using Pillow
img = PILImage.open(png_file_name)
img1 = PILImage.open(png2)

# Convert the Pillow images to openpyxl Image objects
xl_img = XLImage(img)
xl_img1 = XLImage(img1)

# Set the width and height of the images to fit the cell dimensions
xl_img.width = 850
xl_img.height = 550
xl_img1.width = 850
xl_img1.height = 550

# Add image to Sheet1 at cell H1
ws.add_image(xl_img, 'H1')

# Add image to Sheet3 (Sheet2 in your comment) at cell A1
new_ws.add_image(xl_img1, 'A1')

# Save the modified workbook
wb.save(xlfile_name)
